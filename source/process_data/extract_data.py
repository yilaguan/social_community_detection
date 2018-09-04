# coding=utf-8

# Created by JetBrains Pycharm
# @Time      : 2018/9/3 20:18
# @Author    : zhangliang
# @File      : extract_data.py
# @Email     : zhangliangxgd@163.com

import xlrd
import openpyxl
import sys
import math
import pdb
from source.model.syndrome import Syndrome

if sys.getdefaultencoding() != 'utf-8':
    reload(sys)
    sys.setdefaultencoding('utf-8')


# 抽取特殊证素
def extract_special_syndrome_element_data(data_path, syndrome):
    data = xlrd.open_workbook(data_path)
    table = data.sheets()[0]
    nrows = table.nrows
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(title=syndrome_element)
    # 生成第一行数据
    for i in xrange(1, 13):
        worksheet.cell(1, i, table.cell(0, i - 1).value.encode('utf-8'))
    worksheet.cell(1, syndrome.get_column_number(), table.cell(0, syndrome.get_column_number()).value.encode('utf-8'))
    for i in xrange(0, syndrome.get_medicine_end() - syndrome.get_medicine_start() + 1):
        worksheet.cell(1, 14 + i, table.cell(0, syndrome.get_medicine_start() + i - 1).value.encode('utf-8'))
    worksheet_nraws = 2

    # 生成其他行的数据, 抽取出所有该证素为1的行，并且保存到新的sheet中
    for i in xrange(1, nrows):
        if math.fabs(table.cell(i, syndrome.get_column_number()).value - 1.0) < 0.00001:
            for j in xrange(1, 13):
                worksheet.cell(worksheet_nraws, j).value = table.cell(i, j - 1).value
            worksheet.cell(worksheet_nraws, syndrome.get_column_number()).value = table.cell(i, syndrome.get_column_number()).value
            for j in xrange(0, syndrome.get_medicine_end() - syndrome.get_medicine_start() + 1):
                worksheet.cell(worksheet_nraws, 14 + j).value = table.cell(i, syndrome.get_medicine_start() + j - 1).value
            worksheet_nraws += 1

    use_cols = []
    useless_cols = []
    for ncol in xrange(14, worksheet.max_column + 1):
        temp = 0
        for nrow in xrange(2, worksheet.max_row + 1):
            temp += worksheet.cell(nrow, ncol).value
        if math.fabs(temp - 0.0) > syndrome.get_threshold():
            use_cols.append(ncol)
        else:
            useless_cols.append(ncol)

    print "除去前面13列固有列之外的有效列数为： " + str(use_cols.__len__())
    print "除去前面13列固有列之外的有效列数为： " + str(useless_cols.__len__())
    print "其中，无效的列标号为： "
    print useless_cols

    # 删除所有列和为0的无效列
    worksheet_delete_useless_col = workbook.create_sheet(title="delete_useless_cols")
    # 前面13行保留
    for i in xrange(1, worksheet.max_row + 1):
        for j in xrange(1, 14):
            worksheet_delete_useless_col.cell(i, j).value = worksheet.cell(i, j).value
        for j in xrange(0, use_cols.__len__()):
            worksheet_delete_useless_col.cell(i, 14 + j).value = worksheet.cell(i, use_cols[j]).value

    # 找到无效行，即行和为0的行
    use_rows = [1]  # 有效行列表,第一行为药名行，故初始化加入
    useless_rows = []  # 无效行列表
    row_threshold = 0.5  # 设置无效行的阈值
    for row in xrange(2, worksheet_delete_useless_col.max_row + 1):
        temp = 0
        for col in xrange(14, worksheet_delete_useless_col.max_column + 1):
            temp += worksheet_delete_useless_col.cell(row, col).value
        if temp > 0.5:
            use_rows.append(row)
        else:
            useless_rows.append(row)
    print "删除无效列后的数据总行数为： " + str(worksheet_delete_useless_col.max_row)
    print "删除无效行后的数据总行数为： " + str(use_rows.__len__())
    print "其中无效的行如下列表所示： "
    print useless_rows

    # 保存有效行的数据
    worksheet_delete_useless_row = workbook.create_sheet(title="delete_useless_rows")
    for i in xrange(1, use_rows.__len__() + 1):
        for j in xrange(1, worksheet_delete_useless_col.max_column + 1):
            try:
                worksheet_delete_useless_row.cell(i, j).value = worksheet_delete_useless_col.cell(use_rows[i - 1], j).value
            except:
                pdb.set_trace()
    workbook.save(syndrome.get_file_name())

if __name__ == '__main__':
    data_path = "../../data/traditional_chinese_medicine_data.xlsx"
    syndrome_element = "phlegm_with_medicines"
    column_number = 13
    medicine_start = 180
    medicine_end = 554
    threshold = 0.5
    file_name = "phlegm_with_medicines.xlsx"
    syndrome = Syndrome("phlegm_with_medicines", column_number=column_number, medicine_start=medicine_start, medicine_end=medicine_end, threshold=threshold
                        , file_name=file_name)
    extract_special_syndrome_element_data(data_path, syndrome)
