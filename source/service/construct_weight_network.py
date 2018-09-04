# coding=utf-8

# Created by JetBrains Pycharm
# @Time      : 2018/9/4 14:58
# @Author    : zhangliang
# @File      : construct_weight_network.py
# @Email     : zhangliangxgd@163.com
import openpyxl
import time
import math


def construct_weight_network(threshold, data_path, file_name):
    print "开始读取地址为： " + data_path + "的数据"
    workbook = openpyxl.load_workbook(filename=data_path)
    worksheet = workbook.get_sheet_by_name("delete_useless_rows")
    print "正确读取excel的数据"
    medicine_name = []
    for i in xrange(14, worksheet.max_column + 1):
        medicine_name.append(worksheet.cell(1, i).value.encode('utf-8'))
    print "总共的药名数目为： " + str(medicine_name.__len__())
    file_write = open(file_name + ".gml", 'w')
    start_time = time.strftime('%Y-%m-%d %X', time.localtime())
    file_write.write('Creator "Zhang Liang on " ' + start_time + "\n")
    file_write.write('graph\n[\n')
    for i in xrange(len(medicine_name)):
        file_write.write('  node\n  [\n    id ' + str(i + 1) + '\n    label "' + medicine_name[i] + '"\n  ]\n')
    dict_edges = {}
    only_one_medicine_row_id = []
    none_medicine_row_id = []
    for row in xrange(2, worksheet.max_row + 1):
        nodes = []  # 存储当前行中大于1的结点
        for col in xrange(1, worksheet.max_column - 12):  # 需要做位移
            if math.fabs(worksheet.cell(row, col + 13).value - 1.0) < 0.000001:
                nodes.append(col)
        # 将自有一个数据和没有数值的行号进行收集，并在最后告诉用户，是否需要修改
        if len(nodes) == 1 or len(nodes) == 0:
            if len(nodes) == 1:
                only_one_medicine_row_id.append(row)
            else:
                none_medicine_row_id.append(row)
        else:
            for index in xrange(len(nodes)):
                source_node = nodes[index]
                if source_node not in dict_edges:
                    dict_edges[source_node] = {}
                if (index + 1) == (len(nodes) - 1):
                    if nodes[index + 1] not in dict_edges[source_node]:
                        dict_edges[source_node][nodes[index + 1]] = 1
                    else:
                        dict_edges[source_node][nodes[index + 1]] += 1
                    index += 1
                else:
                    for index_next in xrange(index + 1, len(nodes)):
                        target_node = nodes[index_next]
                        if target_node not in dict_edges[source_node]:
                            dict_edges[source_node][target_node] = 1
                        else:
                            dict_edges[source_node][target_node] += 1
    file_write_oslom = open(file_name + "_oslom.txt", 'w')
    num_edges = 0
    num_edges_delete = 0
    all_weight = 0

    for key, value in dict_edges.iteritems():
        for i, j in value.iteritems():
            all_weight += j
            if j > threshold:
                file_write_oslom.write(str(key) + " " + str(i) + " " + str(j) + "\n")
                num_edges += 1
                file_write.write(
                    '  edge\n  [\n    source ' + str(key) + '\n    target ' + str(i) + '\n    weight ' + str(
                        j) + '\n  ]\n')
            else:
                num_edges_delete += 1

    file_write_oslom.close()
    file_write.write(']\n')
    file_write.close()
    print "总的边数为： %d" % num_edges
    print "总的权重为： %d" % all_weight
    print "被删除的边数为： %d" %num_edges_delete


if __name__ == '__main__':
    threshold = 0.5
    data_path = "../process_data/phlegm_with_medicines.xlsx"
    file_name = "../../data/phlegm_weight_network"
    construct_weight_network(threshold=threshold, data_path=data_path, file_name=file_name)
